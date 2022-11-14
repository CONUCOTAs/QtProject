import sys
import sqlite3
import xlsxwriter
import openpyxl
import csv
from os import listdir
from Second_Form import SecondForm

from PyQt5.QtCore import QSize, Qt
from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import QApplication, QLabel, QMainWindow, QFileDialog, QPushButton, QTableWidget, QWidget, \
    QGridLayout, QTableWidgetItem, QLineEdit

SCREEN_SIZE = [1024, 768]


class Example(QMainWindow):
    def __init__(self):
        files = listdir(".")
        mysql = list(filter(lambda x: x.endswith('.sqlite'), files))
        counter = 0
        while f"specCourse_{counter}.sqlite" in mysql:
            counter += 1
        self.nameSQL = f"specCourse_{counter}.sqlite"
        self.isTable = False
        self.search_surname = ''
        self.search_name = ''
        self.id = 1
        self.connection = sqlite3.connect(self.nameSQL)
        self.cursor = self.connection.cursor()
        self.listCourses = ["Робототехника (8-11)",
                            'Электронные приборы (8-10)',
                            '3Д Моделирование (8-10)',
                            'Типографское дело (8-10)',
                            'Программирование игр на Unity (8-9)',
                            'Экономика (8-9)',
                            'Программирование игр на python (8-9)',
                            'Олимпиадная математика (8-11)',
                            'Олимпиадная физика (8-11)',
                            'Олимпиадное программирование (8-11)',
                            'Квадрокоптеры (8-10)',
                            'Робофутбол (8-10)',
                            'Нейроинтерфейсы (8-10)',
                            'Лицей академии Яндекса',
                            'Планиметрия',
                            '3D дизайн в 3Ds MAX',
                            'Выбрано меньше 2 спецкурсов',
                            'Внешнее']
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setMinimumSize(QSize(*SCREEN_SIZE))  # Set sizes
        self.setWindowTitle("Преобразователь")  # Set the window title
        central_widget = QWidget(self)  # Create a central widget
        self.setCentralWidget(central_widget)  # Install the central widget

        self.grid_layout = QGridLayout(self)  # Create QGridLayout
        central_widget.setLayout(self.grid_layout)  # Set this layout in central widget

        self.button_open = QPushButton("Открыть форму", self)
        self.button_open.resize(self.button_open.sizeHint())
        self.button_open.clicked.connect(self.open_table)
        self.grid_layout.addWidget(self.button_open, 0, 0)

        self.pixmap = QPixmap('orig.jpg')
        self.image = QLabel(self)
        self.image.move(80, 60)
        self.image.resize(250, 250)
        self.image.setPixmap(self.pixmap)

    def pre_tables(self):
        self.button_save = QPushButton("Сохранить таблицу", self)
        self.button_save.resize(self.button_save.sizeHint())
        self.button_save.clicked.connect(self.save_table)

        self.table_change = QPushButton("Сохранить изменения", self)
        self.table_change.resize(self.table_change.sizeHint())
        self.table_change.clicked.connect(self.change_table)

        self.search = QPushButton("Поиск", self)
        self.search.resize(self.search.sizeHint())
        self.search.clicked.connect(self.tables)

        self.button_saper = QPushButton("Отдых", self)
        self.button_saper.resize(self.button_saper.sizeHint())
        self.button_saper.clicked.connect(self.games)

        self.grid_layout.removeWidget(self.button_open)

        self.grid_layout.addWidget(self.button_save, 0, 0)
        self.grid_layout.addWidget(self.search, 0, 1)
        self.grid_layout.addWidget(self.table_change, 4, 0)
        self.grid_layout.addWidget(self.button_saper, 5, 0)
        self.button_open.hide()

        self.label_surname = QLabel(self)
        self.label_surname.setText('Поиск по фамилии:')
        self.label_name = QLabel(self)
        self.label_name.setText('Поиск по имени:')
        self.grid_layout.addWidget(self.label_surname, 1, 0)
        self.grid_layout.addWidget(self.label_name, 2, 0)

        self.surname_out = QLineEdit(self)
        self.name_out = QLineEdit(self)
        self.grid_layout.addWidget(self.surname_out, 1, 1)
        self.grid_layout.addWidget(self.name_out, 2, 1)
        self.tables()

    def tables(self):
        if not self.isTable:
            self.table = QTableWidget(self)
            self.isTable = True
        else:
            self.table.clear()
        self.count_row = 0
        self.table.setColumnCount(len(self.listCourses))
        self.result_course = [i[0] for i in self.cursor.execute("""SELECT name FROM Course""").fetchall()]
        # print(result_course)
        self.table.setHorizontalHeaderLabels(self.result_course)

        self.search_surname = self.surname_out.text().lower()
        self.search_name = self.name_out.text().lower()
        for name_course in self.result_course:
            student = self.cursor.execute(f'''SELECT surname, name, class FROM student WHERE id_student IN 
                        (SELECT student_id FROM selection WHERE course_id IN 
                        (SELECT id from Course WHERE name LIKE "{name_course}"))
                        AND surname LIKE '{self.surname_out.text().lower()}%'
                        AND name LIKE '{self.name_out.text().lower()}%' ''').fetchall()
            if len(student) > self.count_row:
                self.count_row = len(student)
                self.table.setRowCount(self.count_row)
            student.sort(key=lambda x: (-int(x[2]), x[0], x[1]))
            for name_student in student:
                self.table.setItem(student.index(name_student), self.result_course.index(name_course), QTableWidgetItem(
                    f"{name_student[0].capitalize()} {name_student[1].capitalize()} {name_student[2]}"))

        self.table.resizeColumnsToContents()

        self.grid_layout.addWidget(self.table, 4, 1)  # Adding the table to the grid

    def save_table(self):
        self.tables()
        with open('results.csv', 'w', newline='', encoding='utf8') as csvfile:
            writer = csv.writer(
                csvfile, delimiter=';', quotechar='"',
                quoting=csv.QUOTE_MINIMAL)
            # Получение списка заголовков
            writer.writerow(
                [self.table.horizontalHeaderItem(i).text()
                 for i in range(self.table.columnCount())])
            for i in range(self.table.rowCount()):
                row = []
                for j in range(self.table.columnCount()):
                    item = self.table.item(i, j)
                    if item is not None:
                        row.append(item.text())
                    else:
                        row.append('')
                writer.writerow(row)
        self.from_csv_to_xlsx()

    def from_csv_to_xlsx(self):
        with open('results.csv', encoding='utf8') as csvfile:
            reader = csv.reader(csvfile, delimiter=';', quotechar='"')
            for index, row in enumerate(reader):
                if index == 0:
                    letters_lenOne = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
                    letters, counter = [], 0
                    counter_own = 0
                    while len(letters) < len(row) * 3:
                        if len(letters_lenOne) > counter_own:
                            letters.append(letters_lenOne[counter])
                        else:
                            letters.append(letters_lenOne[len(letters) // 26 - 1] + letters_lenOne[counter])
                        counter = (counter + 1) % 26
                        counter_own += 1

                    self.workbook = xlsxwriter.Workbook('Спецкурсы.xlsx')
                    self.worksheet = self.workbook.add_worksheet()
                    self.worksheet.set_column(f'{letters[0]}:{letters[-1]}', 20)
                    merge_format = self.workbook.add_format({'align': 'center'})
                    counter = 0

                    for i in row:
                        self.worksheet.merge_range(f'{letters[counter]}1:{letters[counter + 2]}1', i.capitalize(),
                                              merge_format)
                        counter += 3
                else:
                    for el in range(len(row)):
                        if len(row[el]) != 0:
                            self.worksheet.write(f'{letters[el * 3]}{index + 1}', row[el].split()[0])
                            self.worksheet.write(f'{letters[el * 3 + 1]}{index + 1}', row[el].split()[1])
                            self.worksheet.write(f'{letters[el * 3 + 2]}{index + 1}', int(row[el].split()[2]))

            self.workbook.close()

    def change_table(self):
        for name_course in self.result_course:
            i = self.result_course.index(name_course)
            row = []
            for j in range(self.table.rowCount()):
                item = self.table.item(j, i)
                if item is not None:
                    row.append(item.text().lower().split())

            student = [list(i) for i in self.cursor.execute(f'''SELECT surname, name, class FROM student WHERE 
            id_student IN 
                        (SELECT student_id FROM selection WHERE course_id IN 
                        (SELECT id from Course WHERE name LIKE "{name_course}"))
                        AND surname LIKE '{self.surname_out.text().lower()}%'
                        AND name LIKE '{self.name_out.text().lower()}%' ''').fetchall()]
            if len(student) > self.count_row:
                self.count_row = len(student)
                self.table.setRowCount(self.count_row)
            student.sort(key=lambda x: (-int(x[2]), x[0], x[1]))

            for name in student:
                if name not in row:
                    self.cursor.execute(f'''DELETE FROM selection WHERE student_id IN 
                    (SELECT id_student FROM student WHERE 
                    surname LIKE '{name[0]}' AND name LIKE '{name[1]}' AND class = {int(name[2])}) 
                    AND course_id IN (SELECT id from course WHERE name LIKE '{name_course}')''')
                    self.connection.commit()

            for name in row:
                if name not in student and len(name) == 3 and name[-1].isnumeric() and int(name[-1]) == float(name[-1]):
                    result = self.cursor.execute(f'''SELECT id_student FROM student WHERE 
                    surname LIKE '{name[0]}' AND name LIKE '{name[1]}' AND class = {int(name[2])}''').fetchall()
                    if len(result) == 0:
                        print(name)
                        self.cursor.execute(f'''INSERT INTO student VALUES 
                        ({self.id}, '{name[0]}', '{name[1]}', {name[2]})''')
                        result = [[self.id]]
                        self.id += 1
                    self.cursor.execute(f'''INSERT INTO selection VALUES 
                                            ({result[0][0]}, {self.result_course.index(name_course) + 1})''')
                self.connection.commit()

    def open_table(self):
        self.cursor.execute('''CREATE TABLE Course (
                                        id INT PRIMARY KEY,
                                        name VARCHAR(255) NOT NULL
                                        )''')
        for counter in range(len(self.listCourses)):
            self.cursor.execute(f"""INSERT INTO Course VALUES ({counter + 1}, "{self.listCourses[counter]}")""")

        xlsxForm = QFileDialog.getOpenFileName(self, 'Выберите файл с формой', '')[0]
        openFileXlsx = openpyxl.load_workbook(xlsxForm)
        selection = openFileXlsx[openFileXlsx.get_sheet_names()[0]]

        self.cursor.execute('''CREATE TABLE student (
                                        id_student INTEGER PRIMARY KEY AUTOINCREMENT,
                                        surname VARCHAR(255) NOT NULL,
                                        name VARCHAR(255) NOT NULL,
                                        class VARCHAR(255) NOT NULL
                                        )''')

        self.cursor.execute('''CREATE TABLE selection (
                                                student_id INT NOT NULL,
                                                course_id INT NOT NULL,
                                                FOREIGN KEY (student_id) REFERENCES student (id),
                                                FOREIGN KEY (course_id) REFERENCES Course (id)
                                                )''')

        start = 2

        while selection.cell(row=start, column=2).value is not None:
            AllName = selection.cell(row=start, column=2).value.split()
            surname = AllName[0].lower()
            name = AllName[1].lower()
            numberClass = selection.cell(row=start, column=3).value[:-1]
            courses = selection.cell(row=start, column=4).value.split(", ")
            start += 1
            # print(name, surname, numberClass, courses)
            result = self.cursor.execute(f"""SELECT * FROM student
                        WHERE name LIKE '{name}' AND 
                        surname LIKE '{surname}' """).fetchall()
            # print(result)
            if len(result) == 0:
                self.cursor.execute(f'''INSERT INTO student VALUES
                            ({self.id}, "{surname}", "{name}", {numberClass}) ''')
                self.id += 1
            else:
                self.cursor.execute(f"""DELETE FROM selection WHERE student_id IN
                                (SELECT id_student From student
                                WHERE name LIKE '{name}' AND 
                                surname LIKE '{surname}')""")
            result_student = self.cursor.execute(f"""SELECT id_student from student WHERE
                                        name LIKE '{name}' AND surname LIKE '{surname}'""").fetchall()[0][0]
            for course in courses:
                # print(course)
                result_course = self.cursor.execute(f"""SELECT id from Course WHERE
                                            name LIKE '{course}'""").fetchall()
                if len(result_course) == 0:
                    result_course = len(self.listCourses)
                else:
                    result_course = result_course[0][0]
                # print(result_student, result_course)
                self.cursor.execute(f"""INSERT INTO selection(student_id, course_id)
                            VALUES ({result_student}, {result_course})""")
            if len(courses) < 2:
                self.cursor.execute(f"""INSERT INTO selection(student_id, course_id)
                                        VALUES ({result_student}, {len(self.listCourses) - 1})""")
        self.connection.commit()

        with open('course.csv', 'w', newline='', encoding="utf8") as csvfile:
            writer = csv.writer(
                csvfile, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            writer.writerow(['id', 'name'])
            results = self.cursor.execute("""SELECT * from Course""").fetchall()
            for result in results:
                writer.writerow(list(result))
        self.pre_tables()

    def games(self):
        self.second_form = SecondForm()
        self.second_form.show()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = Example()
    ex.show()
    sys.exit(app.exec())
